import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import os
import io
from datetime import datetime
import threading
import time
import cv2
import numpy as np
import serial.tools.list_ports
from config.config_manager import get_config, set_config, save_config
import sys


# 导入RFID和OCR功能
try:
    from rfid_util import rfid_util, EpcFrameDetectedException, TimeoutDetectedException
    RFID_AVAILABLE = True
except ImportError as e:
    print(f"警告: 无法导入RFID工具: {e}")
    RFID_AVAILABLE = False
    # 定义占位异常类
    class EpcFrameDetectedException(Exception):
        pass
    class TimeoutDetectedException(Exception):
        pass

try:
    from rapidocr_onnxruntime import RapidOCR
    OCR_AVAILABLE = True
except ImportError as e:
    print(f"警告: 无法导入OCR工具: {e}")
    OCR_AVAILABLE = False

# ==================== 配置常量 ====================
# TID读取配置
DEFAULT_TID_REQUIRED_COUNT = 5      # TID需要连续读取的次数
DEFAULT_TID_MAX_DURATION = 2        # TID读取最大持续时间(秒)

# OCR识别配置  
DEFAULT_OCR_REQUIRED_COUNT = 3      # OCR需要连续识别的次数
DEFAULT_OCR_MAX_ATTEMPTS_MANUAL = 50    # 手动OCR识别最大尝试次数
DEFAULT_OCR_MAX_ATTEMPTS_AUTO = 20      # 自动OCR识别最大尝试次数

# 界面更新配置
DEFAULT_OCR_SLEEP_INTERVAL = 0.1    # OCR识别间隔时间(秒)
DEFAULT_AUTO_GET_INTERVAL = 1       # 自动获取循环间隔时间(秒)

# 其他时间配置
DEFAULT_RFID_OPERATION_DELAY = 0.5  # RFID操作间隔时间(秒)
DEFAULT_CAMERA_FPS_DELAY = 0.03     # 摄像头帧率延迟(秒, 约30fps)
DEFAULT_ERROR_RETRY_DELAY = 2       # 错误重试延迟时间(秒)
DEFAULT_THREAD_STOP_WAIT = 0.5      # 线程停止等待时间(秒)
DEFAULT_CAMERA_STOP_WAIT = 0.1      # 摄像头停止等待时间(秒)
# ==================== 配置常量结束 ====================

# PyInstaller 打包后获取资源路径的工具函数
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath('.'), relative_path)

class DataRecorderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("RFID-TID识别记录器")
        self.root.geometry("1200x800")

        # Excel文件路径（导出时选择）
        self.excel_file_path = None
        self.current_image_path = None
        self.preview_image = None

        # 摄像头相关
        self.camera = None
        self.camera_running = False
        self.camera_thread = None
        self.current_frame = None

        # OCR相关
        self.ocr_reader = None
        self.last_recognized_text = None
        self.ocr_count = 0

        # RFID相关
        self.rfid_connected = False

        # 自动获取相关
        self.auto_running = False
        self.auto_thread = None

        # 数据存储列表
        self.data_list = []  # 存储读取到的数据
        self.data_set = set()  # 用于去重

        # 状态提示相关
        self.status_message_timer = None  # 状态消息定时器

        # 初始化硬件
        self.init_hardware()

        self.setup_ui()

        # 加载配置并初始化设备列表
        self.load_config()
        self.refresh_ports()
        self.refresh_cameras()

    def init_hardware(self):
        """初始化硬件设备"""
        # 初始化OCR
        if OCR_AVAILABLE:
            try:
                print("正在初始化OCR...")
                # self.ocr_reader = RapidOCR()
                # 修改后（指定模型文件夹路径）
                config_path = resource_path("rapidocr_onnxruntime/config.yaml")
                self.ocr_reader = RapidOCR(config_path=config_path)
                print("✅ OCR初始化成功")
                # import rapidocr_onnxruntime
                # print("OCR模型路径：", rapidocr_onnxruntime.__path__)
            except Exception as e:
                print(f"❌ OCR初始化失败: {e}")
                self.ocr_reader = None

        # 初始化RFID（从配置文件读取端口）
        if RFID_AVAILABLE:
            try:
                rfid_port = get_config('rfid_port', 'COM1')
                print(f"正在连接RFID设备 (端口: {rfid_port})...")
                self.rfid_connected = rfid_util.connect()
                if self.rfid_connected:
                    print(f"✅ RFID设备连接成功 (端口: {rfid_port})")
                else:
                    print(f"❌ RFID设备连接失败 (端口: {rfid_port})")
            except Exception as e:
                print(f"❌ RFID连接失败: {e}")
                self.rfid_connected = False

        # 初始化摄像头（从配置文件读取索引）
        try:
            camera_index = get_config('camera_index', 0)
            print(f"正在初始化摄像头 (索引: {camera_index})...")
            self.camera = cv2.VideoCapture(camera_index)
            if self.camera.isOpened():
                print(f"✅ 摄像头初始化成功 (索引: {camera_index})")
            else:
                print(f"❌ 摄像头初始化失败 (索引: {camera_index})")
                self.camera = None
        except Exception as e:
            print(f"❌ 摄像头初始化失败: {e}")
            self.camera = None

    def setup_ui(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 左侧面板（数据输入）
        left_panel = ttk.Frame(main_frame)
        left_panel.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))

        # 右侧面板（摄像头预览）
        right_panel = ttk.Frame(main_frame)
        right_panel.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 设备配置区域
        config_frame = ttk.LabelFrame(left_panel, text="设备配置", padding="10")
        config_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        # RFID端口配置
        ttk.Label(config_frame, text="RFID端口:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.rfid_port_var = tk.StringVar()
        self.rfid_port_combo = ttk.Combobox(config_frame, textvariable=self.rfid_port_var, width=15, state="readonly")
        self.rfid_port_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=(0, 5))
        ttk.Button(config_frame, text="刷新", command=self.refresh_ports).grid(row=0, column=2, padx=(0, 5), pady=(0, 5))

        # 摄像头配置
        ttk.Label(config_frame, text="摄像头:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        self.camera_var = tk.StringVar()
        self.camera_combo = ttk.Combobox(config_frame, textvariable=self.camera_var, width=15, state="readonly")
        self.camera_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 5), pady=(0, 5))
        ttk.Button(config_frame, text="刷新", command=self.refresh_cameras).grid(row=1, column=2, padx=(0, 5), pady=(0, 5))

        # 高级设置行
        adv_settings_frame = ttk.Frame(config_frame)
        adv_settings_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 5))

        # OCR识别次数
        ttk.Label(adv_settings_frame, text="OCR次数:").pack(side=tk.LEFT, padx=(0, 2))
        self.ocr_count_var = tk.StringVar(value=str(DEFAULT_OCR_REQUIRED_COUNT))
        self.ocr_count_spinbox = ttk.Spinbox(adv_settings_frame, from_=1, to=10, textvariable=self.ocr_count_var, width=5)
        self.ocr_count_spinbox.pack(side=tk.LEFT, padx=(0, 10))

        # TID读取次数
        ttk.Label(adv_settings_frame, text="TID次数:").pack(side=tk.LEFT, padx=(0, 2))
        self.tid_count_var = tk.StringVar(value=str(DEFAULT_TID_REQUIRED_COUNT))
        self.tid_count_spinbox = ttk.Spinbox(adv_settings_frame, from_=1, to=10, textvariable=self.tid_count_var, width=5)
        self.tid_count_spinbox.pack(side=tk.LEFT, padx=(0, 15))

        # RFID重置按钮
        ttk.Button(adv_settings_frame, text="RFID重置", command=self.rfid_reset, width=12).pack(side=tk.LEFT)

        # 配置状态显示
        self.config_status_label = ttk.Label(config_frame, text="配置状态：未加载", foreground="gray")
        self.config_status_label.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=(10, 0))

        # 绑定选择事件
        self.rfid_port_combo.bind('<<ComboboxSelected>>', self.on_rfid_port_changed)
        self.camera_combo.bind('<<ComboboxSelected>>', self.on_camera_changed)
        self.ocr_count_var.trace_add("write", self.on_ocr_count_changed)
        self.tid_count_var.trace_add("write", self.on_tid_count_changed)

        # 数据输入区域
        input_frame = ttk.LabelFrame(left_panel, text="数据输入", padding="10")
        input_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 厂家名称
        ttk.Label(input_frame, text="厂家名称:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.manufacturer_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.manufacturer_var, width=30).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 5))

        # TID输入
        ttk.Label(input_frame, text="TID:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        tid_frame = ttk.Frame(input_frame)
        tid_frame.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(0, 5))

        self.tid_var = tk.StringVar()
        ttk.Entry(tid_frame, textvariable=self.tid_var, width=25).grid(row=0, column=0, sticky=(tk.W, tk.E))
        self.tid_auto_btn = ttk.Button(tid_frame, text="自动获取", command=self.auto_get_tid)
        self.tid_auto_btn.grid(row=0, column=1, padx=(5, 0))
        if not RFID_AVAILABLE or not self.rfid_connected:
            self.tid_auto_btn.config(state="disabled")

        # 标签号输入
        ttk.Label(input_frame, text="标签号:").grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        label_frame = ttk.Frame(input_frame)
        label_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=(0, 5))

        self.label_var = tk.StringVar()
        ttk.Entry(label_frame, textvariable=self.label_var, width=25).grid(row=0, column=0, sticky=(tk.W, tk.E))
        self.label_auto_btn = ttk.Button(label_frame, text="自动获取", command=self.auto_get_label)
        self.label_auto_btn.grid(row=0, column=1, padx=(5, 0))
        if not OCR_AVAILABLE or not self.camera:
            self.label_auto_btn.config(state="disabled")

        # 自动获取控制
        auto_frame = ttk.Frame(input_frame)
        auto_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 10))

        self.auto_btn = ttk.Button(auto_frame, text="开始自动获取", command=self.toggle_auto_get)
        self.auto_btn.grid(row=0, column=0, padx=(0, 10))

        self.status_label = ttk.Label(auto_frame, text="状态：未开始", foreground="gray")
        self.status_label.grid(row=0, column=1, sticky=tk.W)

        # 检查硬件可用性
        if not ((RFID_AVAILABLE and self.rfid_connected) or (OCR_AVAILABLE and self.camera)):
            self.auto_btn.config(state="disabled")
            self.status_label.config(text="状态：硬件不可用", foreground="red")
        
        # 图片选择
        ttk.Label(input_frame, text="图片:").grid(row=4, column=0, sticky=tk.W, pady=(0, 5))
        image_frame = ttk.Frame(input_frame)
        image_frame.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=(0, 5))

        ttk.Button(image_frame, text="选择图片", command=self.select_image).grid(row=0, column=0)
        self.camera_capture_btn = ttk.Button(image_frame, text="摄像头捕获", command=self.capture_from_camera)
        self.camera_capture_btn.grid(row=0, column=1, padx=(5, 0))
        if not self.camera:
            self.camera_capture_btn.config(state="disabled")

        self.image_path_label = ttk.Label(image_frame, text="未选择图片（自动获取时将自动捕获）", foreground="gray")
        self.image_path_label.grid(row=0, column=2, padx=(10, 0), sticky=tk.W)

        # 保存数据按钮
        save_frame = ttk.Frame(input_frame)
        save_frame.grid(row=5, column=0, columnspan=2, pady=(10, 0))

        ttk.Button(save_frame, text="保存数据", command=self.save_single_data, style="Accent.TButton").grid(row=0, column=0, padx=(0, 10))
        ttk.Button(save_frame, text="清空输入", command=self.clear_inputs).grid(row=0, column=1)

        # 数据列表区域
        data_list_frame = ttk.LabelFrame(left_panel, text="读取数据列表", padding="10")
        data_list_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))

        # 创建Treeview显示数据
        columns = ("序号", "厂家名称", "TID", "标签号", "图片来源", "时间")
        self.data_tree = ttk.Treeview(data_list_frame, columns=columns, show="headings", height=8)

        # 设置列标题
        for col in columns:
            self.data_tree.heading(col, text=col)

        # 设置列宽
        self.data_tree.column("序号", width=50)
        self.data_tree.column("厂家名称", width=100)
        self.data_tree.column("TID", width=200)
        self.data_tree.column("标签号", width=50)
        self.data_tree.column("图片来源", width=50)
        self.data_tree.column("时间", width=120)

        # 添加滚动条
        scrollbar = ttk.Scrollbar(data_list_frame, orient="vertical", command=self.data_tree.yview)
        self.data_tree.configure(yscrollcommand=scrollbar.set)

        self.data_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))

        # 绑定双击事件进行编辑
        self.data_tree.bind("<Double-1>", self.on_data_tree_double_click)

        # 批量操作按钮
        batch_frame = ttk.Frame(data_list_frame)
        batch_frame.grid(row=1, column=0, columnspan=2, pady=(10, 0))

        ttk.Button(batch_frame, text="清空列表", command=self.clear_data_list).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(batch_frame, text="删除选中", command=self.delete_selected).grid(row=0, column=1, padx=(5, 5))
        ttk.Button(batch_frame, text="导出到Excel", command=self.export_to_excel, style="Accent.TButton").grid(row=0, column=2, padx=(5, 0))
        
        # 右侧摄像头和预览区域
        # 摄像头预览区域
        camera_frame = ttk.LabelFrame(right_panel, text="摄像头预览", padding="10")
        camera_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))

        self.camera_label = ttk.Label(camera_frame, text="摄像头未启动", anchor="center")
        self.camera_label.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 摄像头控制按钮
        camera_control_frame = ttk.Frame(camera_frame)
        camera_control_frame.grid(row=1, column=0, pady=(10, 0))

        self.camera_start_btn = ttk.Button(camera_control_frame, text="启动摄像头", command=self.start_camera)
        self.camera_start_btn.grid(row=0, column=0, padx=(0, 5))

        self.camera_stop_btn = ttk.Button(camera_control_frame, text="停止摄像头", command=self.stop_camera, state="disabled")
        self.camera_stop_btn.grid(row=0, column=1, padx=(5, 0))

        if not self.camera:
            self.camera_start_btn.config(state="disabled")

        # 图片预览区域
        preview_frame = ttk.LabelFrame(right_panel, text="图片预览", padding="10")
        preview_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.preview_label = ttk.Label(preview_frame, text="未选择图片", anchor="center")
        self.preview_label.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # OCR识别结果显示
        ocr_frame = ttk.LabelFrame(right_panel, text="OCR识别结果", padding="10")
        ocr_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))

        self.ocr_result_label = ttk.Label(ocr_frame, text="未开始识别", anchor="center")
        self.ocr_result_label.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # 配置网格权重
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)

        left_panel.columnconfigure(0, weight=1)
        left_panel.rowconfigure(1, weight=1)
        left_panel.rowconfigure(2, weight=1)

        config_frame.columnconfigure(1, weight=1)

        right_panel.columnconfigure(0, weight=1)
        right_panel.rowconfigure(0, weight=1)
        right_panel.rowconfigure(1, weight=1)

        input_frame.columnconfigure(1, weight=1)
        tid_frame.columnconfigure(0, weight=1)
        label_frame.columnconfigure(0, weight=1)
        auto_frame.columnconfigure(1, weight=1)
        image_frame.columnconfigure(2, weight=1)

        data_list_frame.columnconfigure(0, weight=1)
        data_list_frame.rowconfigure(0, weight=1)

        camera_frame.columnconfigure(0, weight=1)
        camera_frame.rowconfigure(0, weight=1)

        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)

        ocr_frame.columnconfigure(0, weight=1)

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # 添加全局状态提示区域（在窗口底部）
        self.global_status_frame = ttk.Frame(self.root)
        self.global_status_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=10, pady=(0, 10))

        self.global_status_label = ttk.Label(
            self.global_status_frame,
            text="",
            foreground="gray",
            font=("TkDefaultFont", 9)
        )
        self.global_status_label.grid(row=0, column=0, sticky=tk.W)

        # 调整主框架的行配置
        self.root.rowconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=0)

    def rfid_reset(self):
        """RFID重置：停止存盘->读TID"""
        if not RFID_AVAILABLE:
            messagebox.showerror("错误", "RFID模块不可用")
            return

        if not self.rfid_connected:
            messagebox.showerror("错误", "RFID设备未连接，请检查设备连接和端口配置")
            return

        try:
            print("🔄 开始RFID重置操作...")
            self.config_status_label.config(text="配置状态：正在执行RFID重置...", foreground="orange")

            # 步骤1: 停止存盘
            print("📋 步骤1: 停止存盘...")
            stop_response = rfid_util.stop_inventory()
            if stop_response:
                print(f"✅ 停止存盘成功: {stop_response.hex(' ').upper()}")
            else:
                print("⚠️ 停止存盘无响应")

            # 等待一段时间确保停止完成
            import time
            time.sleep(DEFAULT_RFID_OPERATION_DELAY)

            # 步骤2: 读TID
            print("🏷️ 步骤2: 读取TID...")
            tid = rfid_util.read_tid()

            if tid:
                print(f"✅ TID切换成功")


                # 显示成功状态
                self.config_status_label.config(text=f"配置状态：RFID重置成功，TID: {tid[:12]}...", foreground="green")
                self.show_status_message(f"RFID重置成功，读取到TID: {tid}", "success")

            else:
                print("❌ TID读取失败")
                self.config_status_label.config(text="配置状态：RFID重置失败，未读取到TID", foreground="red")
                self.show_status_message("RFID重置失败，未读取到TID", "warning")

            print("🔄 RFID重置操作完成")

        except Exception as e:
            error_msg = f"RFID重置失败: {str(e)}"
            print(f"❌ {error_msg}")
            self.config_status_label.config(text="配置状态：RFID重置异常", foreground="red")
            messagebox.showerror("错误", error_msg)

    def show_status_message(self, message, message_type="info", duration=4000):
        """显示非阻塞状态消息

        Args:
            message: 要显示的消息
            message_type: 消息类型 ("success", "info", "warning", "error")
            duration: 显示持续时间（毫秒）
        """
        # 取消之前的定时器
        if self.status_message_timer:
            self.root.after_cancel(self.status_message_timer)

        # 设置消息颜色
        color_map = {
            "success": "#28a745",  # 绿色
            "info": "#17a2b8",     # 蓝色
            "warning": "#ffc107",  # 黄色
            "error": "#dc3545"     # 红色
        }
        color = color_map.get(message_type, "#6c757d")  # 默认灰色

        # 显示消息
        self.global_status_label.config(text=message, foreground=color)

        # 设置定时器清除消息
        self.status_message_timer = self.root.after(duration, self._clear_status_message)

    def _clear_status_message(self):
        """清除状态消息"""
        self.global_status_label.config(text="", foreground="gray")
        self.status_message_timer = None

    def select_image(self):
        """选择图片文件"""
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("Image files", "*.png *.jpg *.jpeg *.gif *.bmp"),
                ("All files", "*.*")
            ],
            title="选择图片文件"
        )
        
        if file_path:
            try:
                # 验证图片文件
                img = Image.open(file_path)
                img.verify()
                
                self.current_image_path = file_path
                self.image_path_label.config(text=os.path.basename(file_path))
                self.show_image_preview(file_path)
                
            except Exception as e:
                messagebox.showerror("错误", f"无法打开图片文件：{str(e)}")
    
    def show_image_preview(self, image_path):
        """显示图片预览"""
        try:
            img = Image.open(image_path)
            
            # 计算预览尺寸（最大300x300）
            max_size = 300
            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
            
            # 转换为tkinter可用的格式
            self.preview_image = ImageTk.PhotoImage(img)
            self.preview_label.config(image=self.preview_image, text="")
            
        except Exception as e:
            self.preview_label.config(image="", text=f"预览失败：{str(e)}")

    def clear_image_preview(self):
        """清空图片预览"""
        self.preview_label.config(image="", text="未选择图片")
        self.preview_image = None

    def auto_get_tid(self):
        """自动获取TID"""
        if not RFID_AVAILABLE:
            messagebox.showerror("错误", "RFID模块不可用")
            return

        if not self.rfid_connected:
            messagebox.showerror("错误", "RFID设备未连接")
            return

        # 显示读取状态
        self.tid_var.set("正在读取TID...")
        self.tid_auto_btn.config(state="disabled")
        self.root.update()

        def read_tid_thread():
            """TID读取线程"""
            try:
                # 使用计数验证读取TID，需要连续读取指定次数相同TID
                tid = rfid_util.read_tid_with_count_verification(
                    required_count=int(self.tid_count_var.get()),
                    max_duration=DEFAULT_TID_MAX_DURATION
                )

                # 在主线程中更新UI
                self.root.after(0, self._update_tid_result, tid)

            except EpcFrameDetectedException as e:
                print(f"⚠️ 手动TID读取时检测到EPC帧: {e}")
                # 在主线程中显示弹窗
                self.root.after(0, self._show_epc_frame_warning)
                # 同时更新TID读取状态
                self.root.after(0, self._update_tid_result, None)
            except TimeoutDetectedException as e:
                print(f"⚠️ 手动TID读取时检测到超时: {e}")
                # 在主线程中显示超时警告
                self.root.after(0, self._show_timeout_warning, str(e))
                # 同时更新TID读取状态
                self.root.after(0, self._update_tid_result, None)
            except Exception as e:
                self.root.after(0, self._update_tid_error, str(e))

        # 在新线程中执行TID读取
        threading.Thread(target=read_tid_thread, daemon=True).start()

    def _update_tid_result(self, tid):
        """更新TID读取结果"""
        self.tid_auto_btn.config(state="normal")

        if tid:
            # 清除当前输入框内容并更新为自动获取的TID
            self.tid_var.set("")  # 先清除
            self.root.update()    # 刷新界面
            self.tid_var.set(tid) # 再设置新值
            self.show_status_message(f"成功读取TID: {tid}", "success")
        else:
            # 失败时清除输入框
            self.tid_var.set("")
            self.show_status_message("未读取到TID，请检查设备连接和标签位置", "warning")

    def _update_tid_error(self, error_msg):
        """更新TID读取错误"""
        self.tid_auto_btn.config(state="normal")
        # 错误时清除输入框
        self.tid_var.set("")
        messagebox.showerror("错误", f"读取TID失败: {error_msg}")
    
    def auto_get_label(self):
        """自动获取标签号"""
        if not OCR_AVAILABLE:
            messagebox.showerror("错误", "OCR模块不可用")
            return

        if not self.camera or not self.camera.isOpened():
            messagebox.showerror("错误", "摄像头不可用")
            return

        if not self.ocr_reader:
            messagebox.showerror("错误", "OCR引擎未初始化")
            return

        # 显示识别状态
        self.label_var.set("正在识别标签号...")
        self.label_auto_btn.config(state="disabled")
        self.root.update()

        def ocr_recognition_thread():
            """OCR识别线程"""
            try:
                # 连续识别逻辑，需要连续指定次数识别到相同的7位标签
                last_recognized = None
                count = 0
                max_attempts = DEFAULT_OCR_MAX_ATTEMPTS_MANUAL  # 最大尝试次数
                attempts = 0

                while attempts < max_attempts:
                    ret, frame = self.camera.read()
                    if not ret:
                        continue

                    # OCR识别
                    ocr_output, _ = self.ocr_reader(frame)
                    if ocr_output:
                        result = [text for _, text, _ in ocr_output]
                    else:
                        result = []

                    # 过滤7位字母+数字
                    seven_tags = [text for text in result if len(text) == 7]

                    if seven_tags:
                        current_text = seven_tags[0]

                        if current_text == last_recognized:
                            count += 1
                            print(f"OCR识别: {current_text} (第{count}次)")

                            # 更新OCR结果显示
                            self.root.after(0, self._update_ocr_display, current_text, count)

                            if count >= int(self.ocr_count_var.get()):  # 连续指定次数识别到相同文本
                                self.root.after(0, self._update_label_result, current_text)
                                return
                        else:
                            last_recognized = current_text
                            count = 1
                            print(f"OCR识别: {current_text} (第1次)")
                            self.root.after(0, self._update_ocr_display, current_text, count)

                    attempts += 1
                    time.sleep(DEFAULT_OCR_SLEEP_INTERVAL)

                # 超时未识别到稳定结果
                self.root.after(0, self._update_label_timeout)

            except Exception as e:
                self.root.after(0, self._update_label_error, str(e))

        # 在新线程中执行OCR识别
        threading.Thread(target=ocr_recognition_thread, daemon=True).start()

    def _update_ocr_display(self, text, count):
        """更新OCR识别结果显示"""
        self.ocr_result_label.config(text=f"识别到: {text} (第{count}/{self.ocr_count_var.get()}次)")

    def _update_label_result(self, label_text):
        """更新标签号识别结果"""
        self.label_auto_btn.config(state="normal")
        # 清除当前输入框内容并更新为自动识别的标签号
        self.label_var.set("")        # 先清除
        self.root.update()            # 刷新界面
        self.label_var.set(label_text) # 再设置新值
        self.ocr_result_label.config(text=f"识别完成: {label_text}")
        self.show_status_message(f"成功识别标签号: {label_text}", "success")

    def _update_label_timeout(self):
        """更新标签号识别超时"""
        self.label_auto_btn.config(state="normal")
        self.label_var.set("")
        self.ocr_result_label.config(text="识别超时")
        self.show_status_message(f"未能识别到连续{self.ocr_count_var.get()}次相同的7位标签号", "warning")

    def _update_label_error(self, error_msg):
        """更新标签号识别错误"""
        self.label_auto_btn.config(state="normal")
        self.label_var.set("")
        self.ocr_result_label.config(text="识别失败")
        messagebox.showerror("错误", f"标签号识别失败: {error_msg}")

    def start_camera(self):
        """启动摄像头预览"""
        if not self.camera or not self.camera.isOpened():
            messagebox.showerror("错误", "摄像头不可用")
            return

        self.camera_running = True
        self.camera_start_btn.config(state="disabled")
        self.camera_stop_btn.config(state="normal")

        def camera_thread():
            """摄像头线程"""
            while self.camera_running:
                ret, frame = self.camera.read()
                if ret:
                    self.current_frame = frame.copy()

                    # 调整图像大小以适应显示
                    display_frame = cv2.resize(frame, (400, 300))
                    display_frame = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)

                    # 转换为PIL图像
                    pil_image = Image.fromarray(display_frame)
                    photo = ImageTk.PhotoImage(pil_image)

                    # 在主线程中更新显示
                    self.root.after(0, self._update_camera_display, photo)

                time.sleep(DEFAULT_CAMERA_FPS_DELAY)  # 约30fps

        self.camera_thread = threading.Thread(target=camera_thread, daemon=True)
        self.camera_thread.start()

    def stop_camera(self):
        """停止摄像头预览"""
        self.camera_running = False
        self.camera_start_btn.config(state="normal")
        self.camera_stop_btn.config(state="disabled")
        self.camera_label.config(image="", text="摄像头已停止")

    def _update_camera_display(self, photo):
        """更新摄像头显示"""
        self.camera_label.config(image=photo, text="")
        self.camera_label.image = photo  # 保持引用

    def capture_from_camera(self):
        """从摄像头捕获图片"""
        if not self.camera or not self.camera.isOpened():
            messagebox.showerror("错误", "摄像头不可用")
            return

        if not self.camera_running:
            messagebox.showwarning("警告", "请先启动摄像头预览")
            return

        if self.current_frame is None:
            messagebox.showerror("错误", "无法获取摄像头画面")
            return

        try:
            # 将OpenCV图像转换为PIL图像
            frame_rgb = cv2.cvtColor(self.current_frame, cv2.COLOR_BGR2RGB)
            pil_image = Image.fromarray(frame_rgb)

            # 保存到临时文件
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                temp_path = tmp_file.name
                pil_image.save(temp_path, 'PNG')

            self.current_image_path = temp_path
            self.image_path_label.config(text="摄像头捕获")

            # 显示预览
            self.show_image_preview(temp_path)

            self.show_status_message("图片捕获成功", "success")

        except Exception as e:
            messagebox.showerror("错误", f"图片捕获失败: {e}")

    def save_single_data(self):
        """保存单条数据到列表"""
        # 获取输入数据
        manufacturer = self.manufacturer_var.get().strip()
        tid = self.tid_var.get().strip()
        label = self.label_var.get().strip()

        # 验证必填字段
        if not manufacturer:
            messagebox.showerror("错误", "请输入厂家名称！")
            return

        if not tid and not label:
            messagebox.showerror("错误", "请至少输入TID或标签号！")
            return

        # 使用当前选择的图片路径
        image_path = self.current_image_path

        # 添加到数据列表
        self.add_data_to_list(tid if tid else None, label if label else None, image_path)

        # 清空TID和标签号输入框（保留厂家名称）
        self.tid_var.set("")
        self.label_var.set("")

        # 清空图片选择
        self.current_image_path = None
        self.image_path_label.config(text="未选择图片")
        self.clear_image_preview()

        # 显示非阻塞成功提示
        self.show_status_message("数据已添加到列表！", "success")

    def clear_inputs(self):
        """清空输入框（保留厂家名称）"""
        # 只清空TID和标签号，保留厂家名称
        self.tid_var.set("")
        self.label_var.set("")

        # 清空图片选择
        self.current_image_path = None
        self.image_path_label.config(text="未选择图片")
        self.clear_image_preview()

        # 清空OCR识别结果显示
        self.ocr_result_label.config(text="未开始识别")

    def toggle_auto_get(self):
        """切换自动获取状态"""
        if self.auto_running:
            self.stop_auto_get()
        else:
            self.start_auto_get()

    def start_auto_get(self):
        """开始自动获取"""
        if not self.manufacturer_var.get().strip():
            messagebox.showerror("错误", "请先输入厂家名称！")
            return

        # 检查是否有可用的硬件
        has_rfid = RFID_AVAILABLE and self.rfid_connected
        has_camera = self.camera and self.camera.isOpened()

        if not has_rfid and not has_camera:
            messagebox.showerror("错误", "没有可用的硬件设备（RFID或摄像头）！")
            return

        # 如果没有手动选择图片但有摄像头，提示将自动捕获
        if not self.current_image_path and has_camera:
            if not messagebox.askyesno("确认", "未选择图片，将使用摄像头自动捕获图片。是否继续？"):
                return

        self.auto_running = True
        self.auto_btn.config(text="停止自动获取")

        # 显示运行状态
        status_parts = []
        if has_rfid:
            status_parts.append("RFID")
        if has_camera:
            status_parts.append("摄像头")

        status_text = f"状态：正在运行 ({'+'.join(status_parts)})"
        self.status_label.config(text=status_text, foreground="green")

        # 启动自动获取线程
        self.auto_thread = threading.Thread(target=self.auto_get_worker, daemon=True)
        self.auto_thread.start()

    def stop_auto_get(self):
        """停止自动获取"""
        self.auto_running = False
        self.auto_btn.config(text="开始自动获取")
        self.status_label.config(text="状态：已停止", foreground="gray")

    def auto_get_worker(self):
        """自动获取工作线程"""
        while self.auto_running:
            try:
                # 获取TID和标签号
                tid = None
                label = None
                captured_image_path = None

                # 尝试获取TID
                if RFID_AVAILABLE and self.rfid_connected:
                    tid = self.get_tid_sync()

                # 尝试获取标签号
                if OCR_AVAILABLE and self.camera and self.camera.isOpened():
                    label = self.get_label_sync()

                # 如果获取到数据，自动捕获摄像头图片
                if (tid and label) and self.camera and self.camera.isOpened():
                    captured_image_path = self.auto_capture_image()

                # 如果获取到数据，添加到列表
                if tid and label:
                    self.root.after(0, self.add_data_to_list, tid, label, captured_image_path)

                # 等待一段时间再继续
                time.sleep(DEFAULT_AUTO_GET_INTERVAL)

            except Exception as e:
                print(f"自动获取错误: {e}")
                time.sleep(DEFAULT_ERROR_RETRY_DELAY)

    def get_tid_sync(self):
        """同步获取TID"""
        try:
            return rfid_util.read_tid_with_count_verification(
                required_count=int(self.tid_count_var.get()),
                max_duration=DEFAULT_TID_MAX_DURATION
            )
        except EpcFrameDetectedException as e:
            print(f"⚠️ 检测到EPC帧，需要重置RFID设备: {e}")
            # 在主线程中显示弹窗
            self.root.after(0, self._show_epc_frame_warning)
            return None
        except TimeoutDetectedException as e:
            print(f"⚠️ 检测到超时响应: {e}")
            # 在主线程中显示超时警告
            self.root.after(0, self._show_timeout_warning, str(e))
            return None
        except Exception as e:
            print(f"TID读取异常: {e}")
            return None

    def _show_epc_frame_warning(self):
        """显示EPC帧检测警告"""
        import tkinter.messagebox as messagebox
        result = messagebox.askyesno(
            "RFID设备状态异常",
            "检测到EPC帧，设备可能处于EPC模式而非TID模式。\n\n"
            "建议执行RFID重置操作：停止存盘->读TID\n\n"
            "是否立即执行RFID重置？",
            icon="warning"
        )

        if result:
            # 用户选择执行重置
            self.rfid_reset()
    
    def _show_timeout_warning(self, error_message):
        """显示超时警告"""
        # 在全局状态栏显示超时提示
        self.show_status_message(f"RFID通信超时: {error_message}", "error", 6000)
        
        # 在控制台打印详细信息
        print(f"❌ RFID通信超时: {error_message}")
        
        # 更新配置状态标签
        self.config_status_label.config(text="配置状态：RFID通信超时", foreground="red")

    def get_label_sync(self):
        """同步获取标签号"""
        try:
            last_recognized = None
            count = 0
            max_attempts = DEFAULT_OCR_MAX_ATTEMPTS_AUTO
            attempts = 0

            while attempts < max_attempts and self.auto_running:
                ret, frame = self.camera.read()
                if not ret:
                    continue

                ocr_output, _ = self.ocr_reader(frame)
                if ocr_output:
                    result = [text for _, text, _ in ocr_output]
                else:
                    result = []
                seven_tags = [text for text in result if len(text) == 7]

                if seven_tags:
                    current_text = seven_tags[0]

                    if current_text == last_recognized:
                        count += 1
                        print(f"OCR识别: {current_text} (第{count}次)")
                        
                        # 更新OCR结果显示（自动获取模式）
                        self.root.after(0, self._update_ocr_display, current_text, count)
                        
                        if count >= int(self.ocr_count_var.get()):  # 连续指定次数识别到相同文本
                            return current_text
                    else:
                        last_recognized = current_text
                        count = 1
                        print(f"OCR识别: {current_text} (第1次)")
                        
                        # 更新OCR结果显示（自动获取模式）
                        self.root.after(0, self._update_ocr_display, current_text, count)

                attempts += 1
                time.sleep(DEFAULT_OCR_SLEEP_INTERVAL)

            return None
        except:
            return None

    def auto_capture_image(self):
        """自动捕获摄像头图片"""
        try:
            if not self.camera or not self.camera.isOpened():
                return None

            # 读取当前帧
            ret, frame = self.camera.read()
            if not ret:
                return None

            # 创建临时文件保存图片
            import tempfile
            temp_dir = tempfile.gettempdir()
            temp_dir = temp_dir + "\\auto_capture"

            # 确保auto_capture目录存在
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
                print(f"创建自动捕获目录: {temp_dir}")

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 精确到毫秒
            temp_image_path = os.path.join(temp_dir, f"auto_capture_{timestamp}.jpg")

            # 保存图片
            success = cv2.imwrite(temp_image_path, frame)
            if not success:
                print(f"❌ 图片保存失败: {temp_image_path}")
                return None

            print(f"自动捕获图片: {temp_image_path}")
            return temp_image_path

        except Exception as e:
            print(f"自动捕获图片失败: {e}")
            return None

    def add_data_to_list(self, tid, label, image_path=None):
        """添加数据到列表"""
        # 创建唯一标识符用于去重
        data_key = f"{tid or 'N/A'}_{label or 'N/A'}"

        if data_key in self.data_set:
            # 如果是自动捕获的临时图片，需要清理
            if image_path and "auto_capture" in image_path and os.path.exists(image_path):
                try:
                    os.unlink(image_path)
                except:
                    pass
            return  # 数据已存在，跳过

        # 添加到去重集合
        self.data_set.add(data_key)

        # 确定使用的图片路径：优先使用自动捕获的图片，其次使用手动选择的图片
        final_image_path = image_path if image_path else self.current_image_path

        # 创建数据记录
        data_record = {
            'manufacturer': self.manufacturer_var.get().strip(),
            'tid': tid or 'N/A',
            'label': label or 'N/A',
            'image_path': final_image_path,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'auto_captured': bool(image_path and "auto_capture" in image_path)  # 只有自动捕获的图片才标记为自动捕获
        }

        # 添加到数据列表
        self.data_list.append(data_record)

        # 更新界面显示
        self.update_data_tree()

        # 更新图片预览（显示最新保存的图片）
        if final_image_path and os.path.exists(final_image_path):
            self.show_image_preview(final_image_path)
            self.image_path_label.config(text=os.path.basename(final_image_path))

        # 更新OCR识别结果显示
        if label and label != 'N/A':
            self.ocr_result_label.config(text=f"识别完成: {label}")

        # 更新输入框显示最新数据
        if tid and tid != 'N/A':
            self.tid_var.set(tid)
        if label and label != 'N/A':
            self.label_var.set(label)

        # 更新状态
        status_text = f"状态：已获取 {len(self.data_list)} 条数据"
        if image_path:
            status_text += " (含自动捕获图片)"
        self.status_label.config(text=status_text, foreground="blue")

    def update_data_tree(self):
        """更新数据树显示"""
        # 清空现有数据
        for item in self.data_tree.get_children():
            self.data_tree.delete(item)

        # 添加新数据
        for i, data in enumerate(self.data_list, 1):
            # 确定图片来源
            if data.get('auto_captured', False):
                image_source = "自动捕获"
            elif data.get('image_path'):
                image_source = "手动选择"
            else:
                image_source = "无图片"

            self.data_tree.insert("", "end", values=(
                i,
                data['manufacturer'],
                data['tid'],
                data['label'],
                image_source,
                data['timestamp']
            ))

    def clear_data_list(self):
        """清空数据列表"""
        if messagebox.askyesno("确认", "确定要清空所有数据吗？"):
            self.data_list.clear()
            self.data_set.clear()
            self.update_data_tree()
            self.status_label.config(text="状态：列表已清空", foreground="gray")

    def delete_selected(self):
        """删除选中的数据"""
        selected_items = self.data_tree.selection()
        if not selected_items:
            messagebox.showwarning("警告", "请先选择要删除的数据！")
            return

        if messagebox.askyesno("确认", f"确定要删除选中的 {len(selected_items)} 条数据吗？"):
            # 获取选中项的索引
            indices_to_remove = []
            for item in selected_items:
                index = self.data_tree.index(item)
                indices_to_remove.append(index)

            # 按倒序删除，避免索引变化
            for index in sorted(indices_to_remove, reverse=True):
                if 0 <= index < len(self.data_list):
                    # 从去重集合中移除
                    data = self.data_list[index]
                    data_key = f"{data['tid']}_{data['label']}"
                    self.data_set.discard(data_key)

                    # 从列表中删除
                    del self.data_list[index]

            self.update_data_tree()
            self.status_label.config(text=f"状态：剩余 {len(self.data_list)} 条数据", foreground="blue")

    def on_data_tree_double_click(self, event):
        """处理数据列表双击事件"""
        # 获取双击的项目
        item = self.data_tree.selection()[0] if self.data_tree.selection() else None
        if not item:
            return

        # 获取项目的索引
        index = self.data_tree.index(item)
        if 0 <= index < len(self.data_list):
            self.edit_data_item(index)

    def edit_data_item(self, index):
        """编辑数据项"""
        if not (0 <= index < len(self.data_list)):
            return

        data = self.data_list[index]

        # 创建编辑对话框
        edit_dialog = tk.Toplevel(self.root)
        edit_dialog.title("编辑数据")
        edit_dialog.geometry("500x400")
        edit_dialog.resizable(False, False)
        edit_dialog.transient(self.root)
        edit_dialog.grab_set()

        # 居中显示
        edit_dialog.update_idletasks()
        x = (edit_dialog.winfo_screenwidth() // 2) - (edit_dialog.winfo_width() // 2)
        y = (edit_dialog.winfo_screenheight() // 2) - (edit_dialog.winfo_height() // 2)
        edit_dialog.geometry(f"+{x}+{y}")

        # 主框架
        main_frame = ttk.Frame(edit_dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        title_label = ttk.Label(main_frame, text=f"编辑第 {index + 1} 条数据", font=("TkDefaultFont", 12, "bold"))
        title_label.pack(pady=(0, 20))

        # 输入字段
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill=tk.X, pady=(0, 20))

        # 厂家名称
        ttk.Label(fields_frame, text="厂家名称:").grid(row=0, column=0, sticky=tk.W, pady=(0, 10))
        manufacturer_var = tk.StringVar(value=data.get('manufacturer', ''))
        manufacturer_entry = ttk.Entry(fields_frame, textvariable=manufacturer_var, width=40)
        manufacturer_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 10))

        # TID
        ttk.Label(fields_frame, text="TID:").grid(row=1, column=0, sticky=tk.W, pady=(0, 10))
        tid_var = tk.StringVar(value=data.get('tid', ''))
        tid_entry = ttk.Entry(fields_frame, textvariable=tid_var, width=40)
        tid_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 10))

        # 标签号
        ttk.Label(fields_frame, text="标签号:").grid(row=2, column=0, sticky=tk.W, pady=(0, 10))
        label_var = tk.StringVar(value=data.get('label', ''))
        label_entry = ttk.Entry(fields_frame, textvariable=label_var, width=40)
        label_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(0, 10))

        # 图片路径显示
        ttk.Label(fields_frame, text="图片:").grid(row=3, column=0, sticky=tk.W, pady=(0, 10))
        image_path = data.get('image_path', '')
        image_display = os.path.basename(image_path) if image_path else "无图片"
        image_label = ttk.Label(fields_frame, text=image_display, foreground="gray")
        image_label.grid(row=3, column=1, sticky=tk.W, padx=(10, 0), pady=(0, 10))

        # 时间显示
        ttk.Label(fields_frame, text="记录时间:").grid(row=4, column=0, sticky=tk.W, pady=(0, 10))
        time_label = ttk.Label(fields_frame, text=data.get('timestamp', ''), foreground="gray")
        time_label.grid(row=4, column=1, sticky=tk.W, padx=(10, 0), pady=(0, 10))

        # 配置列权重
        fields_frame.columnconfigure(1, weight=1)

        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)

        def save_changes():
            """保存修改"""
            # 获取新值
            new_manufacturer = manufacturer_var.get().strip()
            new_tid = tid_var.get().strip()
            new_label = label_var.get().strip()

            # 验证输入
            if not new_manufacturer:
                messagebox.showerror("错误", "厂家名称不能为空！")
                return

            if not new_tid:
                messagebox.showerror("错误", "TID不能为空！")
                return

            if not new_label:
                messagebox.showerror("错误", "标签号不能为空！")
                return

            # 检查是否有重复数据（排除当前编辑的项）
            new_data_key = f"{new_tid}_{new_label}"
            old_data_key = f"{data['tid']}_{data['label']}"

            if new_data_key != old_data_key and new_data_key in self.data_set:
                messagebox.showerror("错误", "该TID和标签号的组合已存在！")
                return

            # 更新数据
            if old_data_key != new_data_key:
                # 如果数据键发生变化，需要更新去重集合
                self.data_set.discard(old_data_key)
                self.data_set.add(new_data_key)

            # 更新数据记录
            self.data_list[index]['manufacturer'] = new_manufacturer
            self.data_list[index]['tid'] = new_tid
            self.data_list[index]['label'] = new_label

            # 更新界面
            self.update_data_tree()

            # 显示成功消息
            self.show_status_message(f"第 {index + 1} 条数据已更新", "success")

            # 关闭对话框
            edit_dialog.destroy()

        def cancel_edit():
            """取消编辑"""
            edit_dialog.destroy()

        # 按钮
        ttk.Button(button_frame, text="保存", command=save_changes).pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(button_frame, text="取消", command=cancel_edit).pack(side=tk.RIGHT)

    def export_to_excel(self):
        """导出数据到Excel"""
        if not self.data_list:
            messagebox.showwarning("警告", "没有数据可导出！")
            return

        # 选择Excel文件操作
        choice = messagebox.askyesnocancel(
            "选择Excel文件",
            "选择Excel文件操作：\n\n是：新建Excel文件\n否：选择现有Excel文件"
        )

        if choice is None:  # 用户点击取消
            return
        elif choice:  # 用户选择新建
            excel_file = self.create_excel_for_export()
        else:  # 用户选择现有文件
            excel_file = self.select_excel_for_export()

        if not excel_file:
            return

        try:
            # 打开Excel文件
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            success_count = 0

            for data in self.data_list:
                # 找到下一个空行
                next_row = ws.max_row + 1

                # 写入数据
                ws.cell(row=next_row, column=1, value=next_row - 1)  # 序号
                ws.cell(row=next_row, column=2, value=data['manufacturer'])  # 厂家名称
                ws.cell(row=next_row, column=3, value=data['tid'])  # TID
                ws.cell(row=next_row, column=4, value=data['label'])  # 标签号
                ws.cell(row=next_row, column=6, value=data['timestamp'])  # 记录时间

                # 插入图片
                if data['image_path'] and os.path.exists(data['image_path']):
                    try:
                        self.insert_image_to_excel_batch(ws, next_row, 5, data['image_path'])
                    except Exception as e:
                        print(f"插入图片失败: {e}")

                success_count += 1

            # 保存文件
            wb.save(excel_file)
            wb.close()

            self.show_status_message(f"成功导出 {success_count} 条数据到Excel！", "success")

            # 询问是否清空列表
            if messagebox.askyesno("提示", "导出成功！是否清空当前数据列表？"):
                self.clear_data_list()

        except PermissionError as e:
            # 权限错误，通常是文件被其他程序打开
            messagebox.showerror("错误", f"无法保存Excel文件，文件可能已在Excel或其他程序中打开。\n\n请关闭相关程序后重试。\n\n文件路径：{excel_file}")
        except Exception as e:
            # 检查是否是权限相关的错误
            error_msg = str(e).lower()
            if "permission denied" in error_msg or "errno 13" in error_msg:
                messagebox.showerror("错误", f"无法保存Excel文件，文件可能已在Excel或其他程序中打开。\n\n请关闭相关程序后重试。\n\n文件路径：{excel_file}")
            else:
                messagebox.showerror("错误", f"导出失败：{str(e)}")

    def create_excel_for_export(self):
        """为导出创建新的Excel文件"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="创建新Excel文件"
        )

        if file_path:
            try:
                # 创建新的工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "数据记录"

                # 设置表头
                headers = ["序号", "厂家名称", "TID", "标签号", "图片", "记录时间"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                # 设置列宽
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 30
                ws.column_dimensions['F'].width = 20

                wb.save(file_path)
                wb.close()

                self.show_status_message("Excel文件创建成功！", "success")
                return file_path

            except PermissionError as e:
                # 权限错误，通常是文件被其他程序打开
                messagebox.showerror("错误", f"无法创建Excel文件，目标位置可能没有写入权限或文件已被占用。\n\n请检查文件路径权限或选择其他位置。\n\n文件路径：{file_path}")
                return None
            except Exception as e:
                # 检查是否是权限相关的错误
                error_msg = str(e).lower()
                if "permission denied" in error_msg or "errno 13" in error_msg:
                    messagebox.showerror("错误", f"无法创建Excel文件，目标位置可能没有写入权限或文件已被占用。\n\n请检查文件路径权限或选择其他位置。\n\n文件路径：{file_path}")
                else:
                    messagebox.showerror("错误", f"创建Excel文件失败：{str(e)}")
                return None

        return None

    def select_excel_for_export(self):
        """为导出选择现有的Excel文件"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="选择Excel文件"
        )

        if file_path:
            try:
                # 验证文件是否可以打开
                wb = openpyxl.load_workbook(file_path)
                wb.close()

                return file_path

            except PermissionError as e:
                # 权限错误，通常是文件被其他程序打开
                messagebox.showerror("错误", f"无法打开Excel文件，文件可能已在Excel或其他程序中打开。\n\n请关闭相关程序后重试。\n\n文件路径：{file_path}")
                return None
            except Exception as e:
                # 检查是否是权限相关的错误
                error_msg = str(e).lower()
                if "permission denied" in error_msg or "errno 13" in error_msg:
                    messagebox.showerror("错误", f"无法打开Excel文件，文件可能已在Excel或其他程序中打开。\n\n请关闭相关程序后重试。\n\n文件路径：{file_path}")
                else:
                    messagebox.showerror("错误", f"打开Excel文件失败：{str(e)}")
                return None

        return None

    def insert_image_to_excel_batch(self, worksheet, row, col, image_path):
        """批量插入图片到Excel"""
        try:
            # 打开并处理图片
            img = Image.open(image_path)

            # 调整图片大小
            max_width, max_height = 200, 150
            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)

            # 将图片保存到内存中
            img_buffer = io.BytesIO()
            img_format = img.format if img.format else 'PNG'
            img.save(img_buffer, format=img_format)
            img_buffer.seek(0)

            # 创建Excel图片对象
            excel_img = ExcelImage(img_buffer)

            # 设置图片位置
            cell_address = worksheet.cell(row=row, column=col).coordinate
            excel_img.anchor = cell_address

            # 插入图片
            worksheet.add_image(excel_img)

            # 调整行高
            worksheet.row_dimensions[row].height = max_height * 0.75

        except Exception as e:
            raise Exception(f"插入图片失败：{str(e)}")

    def cleanup_resources(self):
        """清理资源"""
        try:
            print("🧹 正在清理资源...")

            # 停止自动获取
            if self.auto_running:
                self.stop_auto_get()
                time.sleep(DEFAULT_THREAD_STOP_WAIT)  # 等待线程结束

            # 停止摄像头
            if self.camera_running:
                self.camera_running = False
                time.sleep(DEFAULT_CAMERA_STOP_WAIT)  # 等待线程结束

            # 释放摄像头
            if self.camera:
                self.camera.release()
                print("✅ 摄像头资源已释放")

            # 关闭RFID连接
            if RFID_AVAILABLE and self.rfid_connected:
                rfid_util.close()
                print("✅ RFID连接已关闭")

            # 清理临时文件
            if self.current_image_path and os.path.exists(self.current_image_path) and "temp" in self.current_image_path:
                try:
                    os.unlink(self.current_image_path)
                    print("✅ 临时文件已清理")
                except:
                    pass

            # 清理自动捕获的临时图片
            self.cleanup_auto_captured_images()

            # 清理状态消息定时器
            if self.status_message_timer:
                self.root.after_cancel(self.status_message_timer)
                self.status_message_timer = None

        except Exception as e:
            print(f"⚠️ 资源清理时发生错误: {e}")

    def cleanup_auto_captured_images(self):
        """清理自动捕获的临时图片"""
        try:
            import tempfile
            temp_dir = tempfile.gettempdir()

            # 清理数据列表中的自动捕获图片
            for data in self.data_list:
                if (data.get('auto_captured', False) and
                    data.get('image_path') and
                    "auto_capture" in data['image_path'] and
                    os.path.exists(data['image_path'])):
                    try:
                        os.unlink(data['image_path'])
                        print(f"✅ 清理自动捕获图片: {os.path.basename(data['image_path'])}")
                    except:
                        pass

            # 清理临时目录中的所有自动捕获图片
            auto_capture_dir = temp_dir + "\\auto_capture"
            try:
                if os.path.exists(auto_capture_dir):
                    for filename in os.listdir(auto_capture_dir):
                        if filename.startswith("auto_capture_") and filename.endswith(".jpg"):
                            file_path = os.path.join(auto_capture_dir, filename)
                            try:
                                os.unlink(file_path)
                                print(f"✅ 清理遗留图片: {filename}")
                            except:
                                pass

                    # 如果目录为空，删除auto_capture目录
                    try:
                        if not os.listdir(auto_capture_dir):
                            os.rmdir(auto_capture_dir)
                            print(f"✅ 清理空目录: {auto_capture_dir}")
                    except:
                        pass
            except:
                pass

        except Exception as e:
            print(f"⚠️ 清理自动捕获图片时发生错误: {e}")

    def load_config(self):
        """加载配置文件"""
        try:
            # 加载RFID端口配置
            rfid_port = get_config('rfid_port', 'COM4')
            self.rfid_port_var.set(rfid_port)

            # 加载摄像头配置
            camera_index = get_config('camera_index', 0)
            self.camera_var.set(f"摄像头 {camera_index}")

            # 加载OCR和TID次数配置
            ocr_count = get_config('ocr_required_count', DEFAULT_OCR_REQUIRED_COUNT)
            self.ocr_count_var.set(str(ocr_count))
            tid_count = get_config('tid_required_count', DEFAULT_TID_REQUIRED_COUNT)
            self.tid_count_var.set(str(tid_count))

            self.config_status_label.config(text="配置状态：已加载", foreground="green")
            print(f"✓ 配置已加载: RFID端口={rfid_port}, 摄像头={camera_index}, OCR次数={ocr_count}, TID次数={tid_count}")

        except Exception as e:
            self.config_status_label.config(text="配置状态：加载失败", foreground="red")
            print(f"⚠️ 配置加载失败: {e}")

    def refresh_ports(self):
        """刷新串口列表"""
        try:
            ports = serial.tools.list_ports.comports()
            port_list = [port.device for port in ports]

            if not port_list:
                port_list = ["无可用端口"]

            self.rfid_port_combo['values'] = port_list

            # 如果当前选择的端口不在列表中，选择第一个
            current_port = self.rfid_port_var.get()
            if current_port not in port_list and port_list[0] != "无可用端口":
                self.rfid_port_var.set(port_list[0])

            print(f"✓ 发现 {len(port_list)} 个串口: {port_list}")

        except Exception as e:
            self.rfid_port_combo['values'] = ["端口检测失败"]
            print(f"⚠️ 串口检测失败: {e}")

    def refresh_cameras(self):
        """刷新摄像头列表"""
        try:
            camera_list = []

            # 检测摄像头（最多检测5个）
            for i in range(5):
                cap = cv2.VideoCapture(i)
                if cap.isOpened():
                    camera_list.append(f"摄像头 {i}")
                    cap.release()

            if not camera_list:
                camera_list = ["无可用摄像头"]

            self.camera_combo['values'] = camera_list

            # 如果当前选择的摄像头不在列表中，选择第一个
            current_camera = self.camera_var.get()
            if current_camera not in camera_list and camera_list[0] != "无可用摄像头":
                print(f"✓ 发现 {len(camera_list)} 个摄像头: {camera_list}")
            self.camera_var.set(camera_list[0])
        except Exception as e:
            self.camera_combo['values'] = ["摄像头检测失败"]
            print(f"⚠️ 摄像头检测失败: {e}")

    def on_rfid_port_changed(self, event):
        """RFID端口选择改变事件"""
        selected_port = self.rfid_port_var.get()
        if selected_port and selected_port != "无可用端口" and selected_port != "端口检测失败":
            try:
                # 保存到配置文件
                set_config('rfid_port', selected_port)
                save_config()

                self.config_status_label.config(text=f"配置状态：RFID端口已保存 ({selected_port})", foreground="blue")
                print(f"✓ RFID端口配置已保存: {selected_port}")

                # 重新初始化RFID连接
                if RFID_AVAILABLE:
                    try:
                        rfid_util.close()
                        self.rfid_connected = rfid_util.connect(selected_port)
                        if self.rfid_connected:
                            print(f"✓ RFID重新连接成功: {selected_port}")
                        else:
                            print(f"⚠️ RFID重新连接失败: {selected_port}")
                    except Exception as e:
                        print(f"⚠️ RFID重新连接异常: {e}")

            except Exception as e:
                self.config_status_label.config(text="配置状态：保存失败", foreground="red")
                print(f"⚠️ RFID端口配置保存失败: {e}")

    def on_ocr_count_changed(self, *args):
        """OCR识别次数改变事件"""
        try:
            count = int(self.ocr_count_var.get())
            set_config('ocr_required_count', count)
            save_config()
            print(f"✓ OCR识别次数已保存: {count}")
        except (ValueError, TypeError):
            # 忽略无效输入
            pass
        except Exception as e:
            print(f"⚠️ OCR识别次数保存失败: {e}")

    def on_tid_count_changed(self, *args):
        """TID读取次数改变事件"""
        try:
            count = int(self.tid_count_var.get())
            set_config('tid_required_count', count)
            save_config()
            print(f"✓ TID读取次数已保存: {count}")
        except (ValueError, TypeError):
            # 忽略无效输入
            pass
        except Exception as e:
            print(f"⚠️ TID读取次数保存失败: {e}")

    def on_camera_changed(self, event):
        """摄像头选择改变事件"""
        selected_camera = self.camera_var.get()
        if selected_camera and selected_camera != "无可用摄像头" and selected_camera != "摄像头检测失败":
            try:
                # 提取摄像头索引
                camera_index = int(selected_camera.split()[-1])

                # 保存到配置文件
                set_config('camera_index', camera_index)
                save_config()

                self.config_status_label.config(text=f"配置状态：摄像头已保存 (索引{camera_index})", foreground="blue")
                print(f"✓ 摄像头配置已保存: 索引{camera_index}")

                # 重新初始化摄像头
                try:
                    if self.camera:
                        self.camera.release()

                    self.camera = cv2.VideoCapture(camera_index)
                    if self.camera.isOpened():
                        print(f"✓ 摄像头重新连接成功: 索引{camera_index}")
                    else:
                        print(f"⚠️ 摄像头重新连接失败: 索引{camera_index}")
                        self.camera = None

                except Exception as e:
                    print(f"⚠️ 摄像头重新连接异常: {e}")
                    self.camera = None

            except Exception as e:
                self.config_status_label.config(text="配置状态：保存失败", foreground="red")
                print(f"⚠️ 摄像头配置保存失败: {e}")


def main():
    """主程序入口"""
    root = tk.Tk()
    app = DataRecorderApp(root)

    # 设置窗口图标（如果有的话）
    try:
        # root.iconbitmap("images\logo.ico")
        icon_path = resource_path(os.path.join("images", "logo.ico"))
        if os.path.exists(icon_path):
            # 确保路径正确且文件存在
            root.iconbitmap(icon_path)
        else:
            print(f"警告：未找到图标文件 {icon_path}")
    except:
        pass  # 如果没有图标文件就忽略

    # 设置窗口关闭事件
    def on_closing():
        if messagebox.askokcancel("退出", "确定要退出程序吗？"):
            # 清理资源
            app.cleanup_resources()
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    # 启动主循环
    root.mainloop()


if __name__ == "__main__":
    main()
 
